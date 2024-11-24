import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import WebDriverWait
import openpyxl
from openpyxl.styles import Alignment


from send_mail import send_mail


current_month = int(datetime.date.today().strftime('%m'))
months = {
    1: '01 - Январь',
    2: '02 - Февраль',
    3: '03 - Март',
    4: '04 - Апрель',
    5: '05 - Май',
    6: '06 - Июнь',
    7: '07 - Июль',
    8: '08 - Август',
    9: '09 - Сентябрь',
    10: '10 - Октябрь',
    11: '11 - Ноябрь',
    12: '12 - Декабрь'
}


options = webdriver.ChromeOptions()
# options.add_argument('--headless')  # работа драйвера в фоновом режиме
driver = webdriver.Chrome(options=options)
url_usd = 'https://www.moex.com/ru/derivatives/currency-rate.aspx?currency=USD_RUB'
url_jpu = 'https://www.moex.com/ru/derivatives/currency-rate.aspx?currency=JPY_RUB'


def get_choice_month():
    time.sleep(2)
    # Выбираем начальную дату -------------------------------------------------------
    driver.find_element(By.XPATH, '//label[@for="fromDate"]').click()
    time.sleep(1)

    # Селектор месяца
    driver.find_element(By.XPATH, '//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[4]/div[1]/div[1]/div[1]').click()
    time.sleep(1)

    # Выбираем месяц
    driver.find_element(By.XPATH, f'//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[2]/div[{current_month-1}]').click()
    time.sleep(1)

    # Выбираем перое число
    driver.find_element(By.XPATH, '//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[4]/div[3]/div[1]/div[contains(text(), "1")]').click()
    time.sleep(1)

    # Выбираем конечную дату -------------------------------------------------------
    driver.find_element(By.XPATH, '//label[@for="tillDate"]').click()
    time.sleep(1)

    # Селектор месяца
    driver.find_element(By.XPATH, '//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[7]/div[1]/div[1]/div[1]').click()
    time.sleep(1)

    # Выбираем месяц
    driver.find_element(By.XPATH, f'//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[5]/div[{current_month-1}]').click()
    time.sleep(1)

    # Выбираем последнее число
    driver.find_element(By.XPATH, '//*[@id="x6yzHnS3PmDg26JWSJJHQUtgRdYhwntYBHnpGe6f7KXJI0S58gvlQbsdGrE0XLam"]/div[7]/div[3]/div[5]/div[last()]').click()
    time.sleep(1)
    # ------------------------------------------------------------------------------
    # Формируем таблицу
    wait.until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), 'Показать')]"))).click()
    time.sleep(1)


def make_text_mail(num):
    if num % 100 in [11, 12, 13, 14]:
        return f'В таблице {num} строк'
    elif num % 10 == 1:
        return f'В таблице {num} строка'
    elif num % 10 in [2, 3, 4]:
        return f'В таблице {num} строки'
    else:
        return f'В таблице {num} строк'


driver.get(url_usd)
wait = WebDriverWait(driver, 20)

wait.until(EC.element_to_be_clickable((By.XPATH, "//*[contains(text(), 'Согласен')]"))).click()


get_choice_month()

# USD_RUB
table = '//*[@id="app"]/div[1]/div[2]/div/div[2]/div[3]/table'

table_1 = wait.until(EC.visibility_of_element_located((By.XPATH, table)))
table_1_value = table_1.text

excel_file = f"result_{datetime.date.today()}.xlsx"

workbook = openpyxl.Workbook()
worksheet = workbook.active

# Задаем заголовки столбцов
worksheet['A1'] = 'Дата USD/RUB'
worksheet['B1'] = 'Курс USD/RUB'
worksheet['C1'] = 'Время USD/RUB'
worksheet['G1'] = 'Результат'

table_rows = table_1_value.replace('\r','').split('\n')
for row in range(7, len(table_rows)):
    table_cols = table_rows[row].split()
    date = table_cols[0]
    value = table_cols[3]
    time_str = table_cols[2]

    worksheet.cell(row=row, column=2, value=value).number_format = '# ##0.0000' + ' ' + u'\u20BD'
    worksheet.cell(row=row, column=5, value=value).number_format = '# ##0.0000' + ' ' + u'\u20BD'
    worksheet.cell(row=row, column=7).number_format = '# ##0.0000' + ' ' + u'\u20BD'

    value_num = float(value)
    worksheet.cell(row=row, column=1, value=date)
    worksheet.cell(row=row, column=2, value=value_num)
    worksheet.cell(row=row, column=3, value=time_str)

workbook.save(excel_file)

# JPY_RUB

driver.get(url_jpu)
time.sleep(5)
get_choice_month()

table_2 = wait.until(EC.visibility_of_element_located((By.XPATH, table)))
table_2_value = table_2.text
driver.close()

workbook = openpyxl.load_workbook(excel_file)
worksheet = workbook.active

worksheet['D1'] = 'Дата JPY/RUB'
worksheet['E1'] = 'Курс JPY/RUB'
worksheet['F1'] = 'Время JPY/RUB'

table_rows = table_2_value.replace('\r','').split('\n')
for row in range(7, len(table_rows)):
    table_cols = table_rows[row].split()
    date = table_cols[0]
    value = table_cols[3]
    time_str = table_cols[2]

    value_num = float(value)
    worksheet.cell(row=row, column=4, value=date)
    worksheet.cell(row=row, column=5, value=value_num)
    worksheet.cell(row=row, column=6, value=time_str)

for row in worksheet.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal='center', vertical='center')

for col in worksheet.columns:
    max_length = 0
    col_letter = col[0].column_letter
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2
    worksheet.column_dimensions[col_letter].width = adjusted_width


worksheet = workbook.active

max_row = worksheet.max_row
for row in range(max_row, 1, -1):
    if not any(cell.value for cell in worksheet[row]):
        worksheet.delete_rows(row)

num_rows = worksheet.max_row
for row in range(2, num_rows + 1):
    worksheet.cell(row=row, column=7).value = f"=B{row}/E{row}"
workbook.save(excel_file)

send_mail(make_text_mail(num_rows), excel_file)
